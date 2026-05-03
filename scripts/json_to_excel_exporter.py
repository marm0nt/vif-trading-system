#!/usr/bin/env python3
"""
JSON to Excel Exporter
Converts VIF analysis JSON to professional Excel spreadsheets
Creates multiple sheets for easy review and filtering
"""

import json
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd

def format_excel_sheet(ws, header_color="#667eea"):
    """Apply professional formatting to worksheet."""
    # Header formatting
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Center align all cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def json_to_excel(json_file):
    """Convert VIF analysis JSON to Excel spreadsheet."""

    if not os.path.exists(json_file):
        print(f"❌ File not found: {json_file}")
        return None

    with open(json_file, 'r') as f:
        data = json.load(f)

    # Get watchlist name and metadata
    watchlist_name = list(data.keys())[0] if data else "analysis"
    watchlist_data = data.get(watchlist_name, {})

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # === Sheet 1: Summary ===
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["VIF Trading Analysis Summary"])
    ws_summary.append([])
    ws_summary.append(["Watchlist", watchlist_data.get("watchlist", "N/A")])
    ws_summary.append(["Analysis Date", watchlist_data.get("analysis_date", "N/A")])
    ws_summary.append(["Tickers Analyzed", watchlist_data.get("tickers_analyzed", 0)])
    ws_summary.append([])
    ws_summary.append(["SIGNALS OVERVIEW"])
    ws_summary.append(["Top 3 BUYs", ", ".join(watchlist_data.get("top_3_buys", [])) or "None"])
    ws_summary.append(["Total Kill Alerts", len(watchlist_data.get("kill_switch_alerts", {}))])

    signals = watchlist_data.get("signals", {})
    buys = len([s for s in signals.values() if s.get("signal") == "BUY"])
    sells = len([s for s in signals.values() if s.get("signal") == "SELL"])
    holds = len([s for s in signals.values() if s.get("signal") == "HOLD"])

    ws_summary.append([])
    ws_summary.append(["Signal Distribution"])
    ws_summary.append(["BUY Signals", buys])
    ws_summary.append(["SELL Signals", sells])
    ws_summary.append(["HOLD Signals", holds])

    format_excel_sheet(ws_summary)

    # === Sheet 2: Kill Switches ===
    ws_kills = wb.create_sheet("Kill Switches", 1)
    kills_data = watchlist_data.get("kill_switch_alerts", {})

    ws_kills.append(["Ticker", "Kill Switch", "Type"])
    for ticker, kill in sorted(kills_data.items()):
        kill_type = kill.split()[0] if isinstance(kill, str) else kill
        ws_kills.append([ticker, kill, kill_type])

    format_excel_sheet(ws_kills)

    # === Sheet 3: All Signals ===
    ws_signals = wb.create_sheet("All Signals", 2)

    signals_list = []
    for ticker, signal_data in sorted(signals.items()):
        signals_list.append({
            "Ticker": ticker,
            "Signal": signal_data.get("signal", ""),
            "Confidence": signal_data.get("confidence", 0),
            "Gamma Regime": signal_data.get("gamma_regime", ""),
            "Volume Signal": signal_data.get("volume_signal", ""),
            "Price": signal_data.get("price", 0),
            "RSI": signal_data.get("rsi", 0),
            "Kill Switch": signal_data.get("kill_switch", ""),
            "Note": signal_data.get("note", "")
        })

    if signals_list:
        df_signals = pd.DataFrame(signals_list)
        for r_idx, row in enumerate(dataframe_to_rows(df_signals, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_signals.cell(row=r_idx, column=c_idx, value=value)

    format_excel_sheet(ws_signals)

    # === Sheet 4: BUY Signals (Filtered) ===
    buy_signals = [s for ticker, s in signals.items() if s.get("signal") == "BUY"]
    if buy_signals:
        ws_buys = wb.create_sheet("BUY Signals", 3)
        buy_list = []
        for ticker, signal_data in signals.items():
            if signal_data.get("signal") == "BUY":
                buy_list.append({
                    "Ticker": ticker,
                    "Confidence": signal_data.get("confidence", 0),
                    "Price": signal_data.get("price", 0),
                    "Volume Signal": signal_data.get("volume_signal", ""),
                    "Note": signal_data.get("note", "")
                })

        if buy_list:
            df_buys = pd.DataFrame(buy_list)
            for r_idx, row in enumerate(dataframe_to_rows(df_buys, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_buys.cell(row=r_idx, column=c_idx, value=value)

        format_excel_sheet(ws_buys)

    # === Sheet 5: SELL Signals (Filtered) ===
    sell_signals = [s for ticker, s in signals.items() if s.get("signal") == "SELL"]
    if sell_signals:
        ws_sells = wb.create_sheet("SELL Signals", 4)
        sell_list = []
        for ticker, signal_data in signals.items():
            if signal_data.get("signal") == "SELL":
                sell_list.append({
                    "Ticker": ticker,
                    "Confidence": signal_data.get("confidence", 0),
                    "Price": signal_data.get("price", 0),
                    "Kill Switch": signal_data.get("kill_switch", ""),
                    "Note": signal_data.get("note", "")
                })

        if sell_list:
            df_sells = pd.DataFrame(sell_list)
            for r_idx, row in enumerate(dataframe_to_rows(df_sells, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_sells.cell(row=r_idx, column=c_idx, value=value)

        format_excel_sheet(ws_sells)

    # === Sheet 6: Kill Switch Analysis ===
    ws_kill_analysis = wb.create_sheet("Kill Switch Analysis", 5)

    kill_types = {}
    for ticker, kill in kills_data.items():
        kill_code = kill.split()[0] if isinstance(kill, str) else kill
        if kill_code not in kill_types:
            kill_types[kill_code] = []
        kill_types[kill_code].append(ticker)

    ws_kill_analysis.append(["Kill Switch Type", "Count", "Affected Tickers"])
    kill_descriptions = {
        "K1": "RSI extreme (>80 or <20)",
        "K2": "High volatility (5d-range>12%)",
        "K3": "Low liquidity (vol<500k)",
        "K6": "Technical breakdown (price<MA20 & vol_weak)"
    }

    for kill_code in sorted(kill_types.keys()):
        tickers = kill_types[kill_code]
        description = kill_descriptions.get(kill_code, kill_code)
        ws_kill_analysis.append([
            f"{kill_code}: {description}",
            len(tickers),
            ", ".join(sorted(tickers))
        ])

    format_excel_sheet(ws_kill_analysis)

    # Save workbook – use JSON timestamp to avoid duplicates on same-JSON re-export
    # Extract timestamp from source JSON filename if available (e.g., analysis_20260502_142605.json)
    json_basename = Path(json_file).stem
    json_ts = ""
    if "_" in json_basename:
        parts = json_basename.split("_")
        if len(parts) >= 2 and parts[-2].isdigit() and parts[-1].isdigit():
            json_ts = f"{parts[-2]}_{parts[-1]}"

    if not json_ts:
        json_ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_file = Path("reports") / f"{watchlist_name.upper()}_ANALYSIS_{json_ts}.xlsx"

    # Skip if file already exists (already exported this JSON)
    if output_file.exists():
        print(f"⊘ Excel already exists: {output_file.name}")
        return str(output_file)

    wb.save(output_file)

    print(f"✅ Excel spreadsheet created: {output_file}")
    return str(output_file)

def export_latest_analysis():
    """Find and export the latest analysis JSON file."""
    reports_dir = Path("reports")
    analysis_files = sorted(reports_dir.glob("analysis_*.json"), reverse=True)

    if analysis_files:
        latest_file = analysis_files[0]
        print(f"📊 Exporting: {latest_file.name}")
        return json_to_excel(str(latest_file))
    else:
        print("❌ No analysis files found in reports/")
        return None

if __name__ == "__main__":
    export_latest_analysis()
